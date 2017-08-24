import codecs
import csv
import json
import re
import warnings
import codecs
from openpyxl import load_workbook, Workbook
from time import strftime

from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.datastructures import MultiValueDictKeyError
from django.views import View
from django.views.generic import DeleteView, DetailView, ListView, TemplateView

from .constants import *
from .forms import FieldSelectorForm, UploadFileForm
from .mixins import CSVResponseMixin
from .models import *
from crm.models import Person, Changes
from crm.views import add_change_record
from crm.constants import GEO_CHOICES, CAT_CHOICES, DIV_CHOICES

######################
# Main page views
######################
class Index(CSVResponseMixin, ListView):
    template_name = 'marketing/index.html'
    context_object_name = 'records'
    queryset = Person.objects.all()
    paginate_by=250
    query_param_terms = {
        'main_category': 'main_category',
        'main_category2': 'main_category2',
        'geo': 'geo',
        'division1': 'division1',
        'division2': 'division2',
        'company': 'company__icontains',
        'name': 'name__icontains',
        'title': 'title__icontains',
        'dept': 'dept__icontains',
        'phone': 'phone__icontains',
        'do_not_call': 'do_not_call',
        'email': 'email__iccontains',
        'do_not_email': 'do_not_email',
        'industry': 'industry__icontains',
        'email_alternate': 'email_alternate__icontains'}

    def _generate_pagination_list(self, context):
        paginator = context['paginator']
        page_obj = context['page_obj']
        page_num = page_obj.number
        last_page = paginator.num_pages
        pagination_list = [1,]

        # pre-page numbers
        if page_num > 500:
            start_num = page_num % 500
            start_num += 500 if start_num == 0 else 0
            while start_num < page_num:
                pagination_list.append(start_num)
                start_num += 500
            pagination_list.append(page_num - 100)
            pagination_list.append(page_num - 10)
        elif page_num > 100:
            start_num = page_num % 100
            start_num += 100 if start_num == 0 else 0
            while start_num < page_num:
                pagination_list.append(start_num)
                start_num += 100
            pagination_list.append(page_num - 10)
        elif page_num > 50:
            pagination_list.append(page_num-50)
            pagination_list.append(page_num - 10)
        elif page_num > 10:
            start_num = page_num % 10
            start_num += 10 if start_num == 0 else 0
            while start_num < page_num:
                pagination_list.append(start_num)
                start_num += 10
        for i in range (page_num-5, page_num):
            if i > 1:
                pagination_list.append(i)

        # post-page numbers
        for i in range(page_num + 1, page_num + 6):
            if i < last_page:
                pagination_list.append(i)
        if last_page - page_num > 500:
            pagination_list.append(page_num + 10)
            pagination_list.append(page_num + 100)
            start_num = page_num + 500
            while start_num < last_page:
                pagination_list.append(start_num)
                start_num += 500
        elif last_page - page_num > 100:
            pagination_list.append(page_num + 10)
            start_num = page_num + 100
            while start_num < last_page:
                pagination_list.append(start_num)
                start_num += 100
        elif last_page - page_num > 50:
            pagination_list.append(page_num + 10)
            pagination_list.append(page_num + 50)
        elif last_page - page_num > 10:
            start_num = page_num + 10
            while start_num < last_page:
                pagination_list.append(start_num)
                start_num += 10

        if last_page > 1:
            pagination_list.append(last_page)

        return pagination_list

    def get_ordering(self):
        """
        Return the field or fields to use for ordering the queryset.
        Overrides default method
        """
        self.order = self.request.GET.get('order', 'asc')
        sort_by = self.request.GET.get('sort_by', None)
        if sort_by and self.order == 'desc':
            sort_by = '-' + sort_by
        return sort_by

    def get_paginate_by(self, queryset):
        return super(Index, self).get_paginate_by(queryset)

    def _filter_queryset(self, queryset):
        query_string = ''
        query_params = {}
        query_prefill = {}
        for param in self.request.GET:
            if param in self.query_param_terms:
                query_string += param + '=' + self.request.GET[param] + '&'
                query_prefill[param] = self.request.GET[param]
                if self.request.GET[param] in ('true', 'false'):
                    tf_bool = self.request.GET[param] == 'true'
                    query_params[self.query_param_terms[param]] = tf_bool
                else:
                    query_params[self.query_param_terms[param]] = \
                        self.request.GET[param]
        query_string = re.sub(r'\s', '%20', query_string)
        query_string = query_string[:-1]
        self.filter_string = query_string if len(query_string) > 0 else None
        queryset = queryset.filter(**query_params)
        self.query_prefill = query_prefill
        return queryset

    def get_queryset(self):
        queryset = super(Index, self).get_queryset()
        queryset = self.filtered_queryset = self._filter_queryset(queryset)
        return queryset

    def paginate_queryset(self, queryset, page_size):
        """
        Paginate the queryset, if needed.
        Override default method to go to first or last page if out of bounds
        """
        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        if page_number < 1:
            page_number = 1
        if page_number > paginator.num_pages:
            page_number = paginator.num_pages
        try:
            page = paginator.page(page_number)
            return (paginator, page, page.object_list, page.has_other_pages())
        except InvalidPage as e:
            raise Http404(_('Invalid page (%(page_number)s): %(message)s') % {
                'page_number': page_number,
                'message': str(e)
            })

    def get_context_data(self, **kwargs):
        context = super(Index, self).get_context_data(**kwargs)
        context['geo_choices'] = GEO_CHOICES
        context['cat_choices'] = CAT_CHOICES
        context['div_choices'] = DIV_CHOICES
        if context['is_paginated']:
            context['pagination_list'] = self._generate_pagination_list(context)
        context['order'] = self.order
        sort_by = self.get_ordering()
        if sort_by and sort_by[0] == '-':
            sort_by = sort_by[1:]
        context['sort_by'] = sort_by
        context['filter_string'] = self.filter_string
        context['query_prefill'] = self.query_prefill
        context['upload_file_form'] = UploadFileForm()
        return context


class UploadFile(TemplateView):
    template_name = 'marketing/upload.html'
    error_message = None
    uploaded_file = None

    def _csv_row_is_not_blank(self, row):
        for cell_num in range(self.num_cols):
            if row[cell_num] not in ('', None):
                return True
        return False

    def _add_csv_row_to_db(self, row, is_first, row_number):
        new_row = UploadedRow(
            parent_file = self.uploaded_file,
            row_is_first=is_first,
            row_number=row_number,
        )
        new_row.save()
        for cell_num in range(self.num_cols):
            new_cell = UploadedCell(
                parent_row = new_row,
                cell_order = cell_num,
                content=row[cell_num],
            )
            new_cell.save()

    def _add_csv_file_to_db(self, decoder):
        f = codecs.iterdecode(
            self.upload_file_form.cleaned_data['marketing_file'],
            decoder
        )
        reader = csv.reader(f)
        if not self.uploaded_file:
            new_file = UploadedFile(
                filename=self.upload_file_form.cleaned_data['marketing_file'].name,
                uploaded_by=self.request.user,
                num_columns=0,
            )
            new_file.save()
            self.uploaded_file = new_file
        is_first_row = True
        self.num_cols = None
        row_number = 0
        for row in reader:
            if not self.num_cols:
                self.num_cols = len(row)
            if self._csv_row_is_not_blank(row):
                self._add_csv_row_to_db(row, is_first_row, row_number)
            is_first_row = False
            row_number += 1
        if self.num_cols:
            self.uploaded_file.num_columns = self.num_cols
            self.uploaded_file.save()

    def _process_csv(self):
        decoder_list = ['utf-8', 'windows-1252']
        if self.request.encoding and self.request_encoding not in decoder_list:
            decoder.insert(0, self.request.encoding)
        successful_transcription = False
        for decode_attempt in range(len(decoder_list)):
            if not successful_transcription:
                try:
                    self._add_csv_file_to_db(decoder_list[decode_attempt])
                    successful_transcription = True
                except UnicodeDecodeError:
                    if self.uploaded_file:
                        UploadedRow.objects.filter(
                            parent_file=self.uploaded_file
                        ).delete()

    def _xlsx_row_is_not_blank(self, ws, row_num, num_cols):
        for col_num in range(1, num_cols+1):
            if ws.cell(row=row_num, column=col_num).value not in ('', None):
                return True
        return False

    def _add_xlsx_row_to_db(self, ws, row_num, num_cols):
        new_row = UploadedRow(
            parent_file = self.uploaded_file,
            row_is_first = row_num == 1,
            row_number = row_num,
        )
        new_row.save()
        for col_num in range(1, num_cols+1):
            new_cell = UploadedCell(
                parent_row=new_row,
                cell_order=col_num,
                content=ws.cell(row=row_num, column=col_num).value
            )
            new_cell.save()

    def _process_xlsx(self, datafile):
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            wb = load_workbook(datafile, read_only=True, data_only=True)
        ws = wb.active
        num_rows = ws.max_row
        num_cols = ws.max_column
        new_file = UploadedFile(
            filename=self.upload_file_form.cleaned_data['marketing_file'].name,
            uploaded_by=self.request.user,
            num_columns=num_cols
        )
        new_file.save()
        self.uploaded_file = new_file
        for row_num in range(1, num_rows+1):
            if self._xlsx_row_is_not_blank(ws, row_num, num_cols):
                self._add_xlsx_row_to_db(ws, row_num, num_cols)

    def post(self, request, *args, **kwargs):
        self.upload_file_form = UploadFileForm(request.POST, request.FILES)
        if self.upload_file_form.is_valid():
            uploaded_file = request.FILES['marketing_file']
            try:
                self._process_xlsx(uploaded_file)
            except Exception as e:
                if self.uploaded_file:
                    self.uploaded_file.delete()
                try:
                    self._process_csv()
                except Exception as e:
                    self.error_message = 'File was not readable.\n' + \
                        'It should be either xlsx or csv format, with ' + \
                        'either utf-8 or windows-1252 encoding.\n' + \
                        'If you are not able to fix this, please talk to ' + \
                        'Chris.\n\n' + \
                        'The specific error message was: \n\n' + str(e)
        else:
            self.error_message = 'Invalid File Submitted'

        datafile = None
        datafile_type_is = None

        context = self.get_context_data(**kwargs)
        return super(UploadFile, self).render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super(UploadFile, self).get_context_data(**kwargs)
        context['error_message'] = self.error_message
        context['unprocessed_files'] = \
            UploadedFile.objects.all().order_by('-uploaded_at')
        context['upload_file_form'] = UploadFileForm()
        return context


#####################
# Ajax views
#####################
class Add(TemplateView):
    template_name = 'marketing/index_addins/table_row.html'

    def _duplicate_person(self, request):
        self._person = Person.objects.get(pk=request.POST['person_id'])
        self._person.pk = None
        self._person.created_by = request.user
        self._person.modified_by = request.user
        self._person.date_created = timezone.now()
        self._person.date_modified = timezone.now()
        self._person.save()

    def _new_person(self, request):
        self._person = Person(
            date_created=timezone.now(),
            created_by=request.user,
            date_modified=timezone.now(),
            modified_by=request.user,
        )
        self._person.save()

    def get(self, request, *args, **kwargs):
        raise Http404()

    def post(self, request, *args, **kwargs):
        if 'person_id' in request.POST:
            try:
                self._duplicate_person(request)
            except Person.DoesNotExist:
                self._new_person(request)
        else:
            self._new_person(request)
        context = self.get_context_data(**kwargs)
        return super(Add, self).render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super(Add, self).get_context_data(**kwargs)
        context['record'] = self._person
        context['geo_choices'] = GEO_CHOICES
        context['cat_choices'] = CAT_CHOICES
        context['div_choices'] = DIV_CHOICES
        return context


class BulkUpdate(View):

    def get(self, request, *args, **kwargs):
        raise Http404()

    def post(self, request, *arts, **kwargs):
        data = json.loads(request.POST['json'])
        person_ids = data['record_list']
        field_data = data['field_dict']
        successful_updates = {}
        Person.objects.filter(id__in=person_ids).update(**field_data)
        return HttpResponse(status=204)


class DeletePerson(View):

    def get(self, request, *args, **kwargs):
        raise Http404()

    def post(self, request, *args, **kwargs):
        person = get_object_or_404(Person, pk=request.POST['record_id'])
        pk = person.pk
        Changes.objects.filter(orig_id=pk).delete()
        if person.has_registration_history():
            add_change_record(person, 'delete')
        person.delete()
        return HttpResponse(status=204)


class DeleteUpload(DeleteView):
    queryset = UploadedFile.objects.all()

    def get(self, request, *args, **kwargs):
        raise Http404()

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return HttpResponse(status=200)


class DownloadErrors(View):

    def _build_row_list(self, row):
        row_text = []
        for i in range(self.upload.num_columns + 1):
            try:
                row_text.append(
                    UploadedCell.objects.get(parent_row=row,
                                             cell_order=i).content
                )
            except UploadedCell.DoesNotExist:
                if i == self.upload.num_columns:
                    row_text.append(row.error_message)
                else:
                    row_text.append('')
        return row_text

    def _csv_response(self, filename):
        filename += 'csv'
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = \
            ('attachment; filename="%s.csv"' % filename)
        writer = csv.writer(response)
        filerows = UploadedRow.objects.filter(parent_file=self.upload)
        for row in filerows:
            row_text = self._build_row_list(row)
            writer.writerow(row_text)
        self.upload.delete()
        return response

    def _xlsx_response(self, filename):
        filename += 'xlsx'
        response = HttpResponse(content_type='application/vnd.mx-excel')
        response['Content-Disposition'] = \
            ('attachment; filename="%s.xlsx"' % filename)
        wb = Workbook()
        ws = wb.active
        filerows = UploadedRow.objects.filter(parent_file=self.upload)
        for row in filerows:
            row_text = self._build_row_list(row)
            ws.append(row)
        wb.save(response)
        self.upload.delete()
        return response

    def get(self, request, *args, **kwargs):
        ok_to_render = False
        self.filetype = request.GET.get('fileformat', None)
        self.upload = get_object_or_404(UploadedFile,
                                        pk=request.GET['file_id'])
        if self.filetype:
            return self.render_to_response()
        else:
            raise Http404('invalid format specified')

    def render_to_response(self):
        if self.upload.filename[-3:] == 'csv':
            filename = self.upload.filename[:-4] + '_errors.'
        else:
            filename = self.upload.filename[:-5] + '_errors.'
        if self.filetype == 'csv':
            return self._csv_response(filename)
        else:
            return self._xlsx_response(filename)


class DownloadList(View):

    def get(self, request, *args, **kwargs):
        pass


class FieldMatcher(DetailView):
    template_name = 'marketing/upload_addins/field_matcher.html'
    queryset = UploadedFile.objects.all()
    context_object_name = 'uploaded_file'

    def _first_ten_rows(self):
        rows = UploadedRow.objects.filter(
            parent_file=self.object,
            row_is_first=False
        ).order_by('row_number')[:10]
        return []

    def get_context_data(self, **kwargs):
        context = super(FieldMatcher, self).get_context_data(**kwargs)
        try:
            first_row = UploadedRow.objects.get(
                parent_file=self.object, row_is_first=True
            )
        except UploadedRow.DoesNotExist:
            first_row = None
        if first_row:
            header_cells = UploadedCell.objects.filter(
                parent_row=first_row
            ).order_by('cell_order')
        else:
            header_cells = None
        context['header_cells'] = header_cells
        context['first_ten_rows'] = UploadedRow.objects.filter(
            parent_file=self.object,
            row_is_first=False,
        ).order_by('row_number')[:10]
        context['selector_form'] = FieldSelectorForm()
        return context


class ProcessUpload(View):

    def get(self, request, *args, **kwargs):
        raise Http404()

    def _delete_file(self):
        num_rows = UploadedRow.objects.filter(parent_file=self.upload).count()
        if num_rows == 0:
            self.upload.delete()
            return True
        elif num_rows == 1:
            row_is_do_not_import = UploadedRow.objects.all()[0].row_is_first
            if row_is_do_not_import:
                self.upload.delete()
                return True
        else:
            return False

    def _import_row(self, row, cell_map):
        person_attrs = {}
        for cell in UploadedCell.objects.filter(parent_row=row):
            if cell.cell_order in cell_map:
                person_attrs[cell_map[cell.cell_order]] = cell.content
        person_id = person_attrs.pop('id', None)
        email = person_attrs['email'] if 'email' in person_attrs else None
        try:
            person = Person.objects.get(pk=person_id)
        except Person.DoesNotExist:
            if email:
                try:
                    person = Person.objects.filter(email=email)[0]
                except IndexError:
                    person = Person(
                        date_created = timezone.now(),
                        created_by = self.request.user,
                    )
        except ValueError:
            row.has_error = True
            row.error_message = str(person_id) + ' is not a valid record id. ' + \
                'It should be an integer.'
            row.save()
            self.rows_imported['total'] += 1
            return
        person_attrs['date_modified'] = timezone.now()
        person_attrs['modified_by'] = self.request.user
        if 'f1-category-split' in person_attrs:
            f1 = person_attrs.pop('f1-category-split')
            try:
                person_attrs['main_category'], \
                    person_attrs['main_category2'] = F1_SPLIT[f1].lower()
            except KeyError:
                person_attrs['main_category'] = \
                    person_attrs['main_category2'] = 'NA'
        if 'geo' in person_attrs:
            person_attrs['geo'] = GEO_DICT[person_attrs['geo'].lower()] \
                if person_attrs['geo'].lower() in GEO_DICT else 'Unknown'

        try:
            for key, value in person_attrs.items():
                setattr(person, key, value)
            person.save()
            self.rows_imported['success'] += 1
            self.rows_imported['total'] += 1
            if not row.row_is_first:
                row.delete()
        except Exception as e:
            row.has_error = True
            row.error_message = str(e)
            row.save()
            self.rows_imported['total'] += 1

    def post(self, request, *arts, **kwargs):
        self.rows_imported = {'success': 0, 'total': 0}
        data = json.loads(request.POST['json'])
        include_first_row = request.POST['ignore_first_row'] == 'false'
        self.upload = UploadedFile.objects.get(pk = data['file_id'])
        cell_map = {int(y[0]):x for x,y in data['field_matches'].items()}
        for row in UploadedRow.objects.filter(parent_file=self.upload):
            if include_first_row or (not row.row_is_first):
                self._import_row(row, cell_map)
        file_fully_processed = self._delete_file()
        response_json = {
            'processComplete': file_fully_processed,
            'rowsImported': self.rows_imported
        }
        return JsonResponse(response_json)


class UpdatePerson(View):

    def get(self, request, *args, **kwargs):
        raise Http404()

    def post(self, request, *args, **kwargs):
        person = get_object_or_404(Person, pk=request.POST['record_id'])
        update_field = request.POST['field']
        new_value = request.POST['new_value']
        old_value = getattr(person, update_field)
        if new_value in ('true', 'false') and old_value in (True, False):
            new_value = new_value == 'true'
        if new_value != old_value:
            setattr(person, update_field, new_value)
            person.date_modified = timezone.now()
            person.modified_by = request.user
            person.save()
        person_vals = {
            'date_modified': person.date_modified.strftime('%m/%d/%Y'),
            'state_prov': person.state_prov(),
        }
        return JsonResponse(person_vals)
