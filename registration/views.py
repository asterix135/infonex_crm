import csv
import datetime
from io import BytesIO
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from string import ascii_uppercase

from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponseRedirect, HttpResponse, Http404, \
    JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import RequestContext
from django.urls import reverse
from django.utils import timezone
from django.utils.datastructures import MultiValueDictKeyError

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate

from .forms import *
from .models import *
from .pdfs import *
from crm.models import Person, Event
from delegate.constants import *
from infonex_crm.settings import BASE_DIR


########################
# HELPER FUNCTIONS
########################
def create_temp_conf_number():
    """ helper function for add_event_option """
    base_number = "TEMP"
    counter = 0
    while True:
        event_number = base_number + str(counter)
        if Event.objects.filter(number=event_number).count() == 0:
            return event_number
        counter += 1


def format_sales_report_header_row(ws, row_num=1):
    """
    Sets first 26 columns of specified row in worksheet to bold
    """
    for letter in ascii_uppercase:
        cell = ws[letter + str(row_num)]
        cell.font = Font(bold=True)


def sales_report_row(regdetail):
    """
    takes a RegDetail object as param
    returns list of values to be inserted into sales report spreadsheet
    """
    registrant = regdetail.registrant
    try:
        invoice_num = regdetail.invoice.pk
        pre_tax_price = regdetail.invoice.pre_tax_price
        pay_date = regdetail.invoice.payment_date
        pay_method = regdetail.invoice.get_payment_method_display()
        fx_rate = regdetail.invoice.fx_conversion_rate
        if regdetail.invoice.sales_credit.first_name or \
            regdetail.invoice.sales_credit.last_name:
            sales_credit = regdetail.invoice.sales_credit.first_name + ' ' + \
                regdetail.invoice.sales_credit.last_name
        else:
            sales_credit = regdetail.invoice.sales_credit.username

    except Invoice.DoesNotExist:
        invoice_num = ''
        pre_tax_price = 0
        pay_date = ''
        pay_method = ''
        fx_rate = 1.0
        sales_credit = ''

    return [sales_credit,
            invoice_num,
            regdetail.conference.number,
            regdetail.conference.title,
            regdetail.get_registration_status_display(),
            registrant.first_name + ' ' + registrant.last_name,
            registrant.company.name,
            regdetail.register_date,
            pay_date,
            pay_method,
            regdetail.cancellation_date,
            pre_tax_price,
            regdetail.conference.billing_currency,
            fx_rate,
            pre_tax_price * fx_rate]


########################
# BASE VIEWS
########################
@login_required
def add_edit_conference(request):
    """ Renders conference page """
    edit_action = 'blank'
    event = None
    event_option_set = None
    conference_edit_form = ConferenceEditForm()
    conference_option_form = ConferenceOptionForm()
    if request.method == 'GET':
        edit_action = request.GET.get('action', 'blank')
        if edit_action not in ('blank', 'new', 'edit'):
            edit_action = 'blank'
        if edit_action == 'edit':
            try:
                event = Event.objects.get(pk=request.GET['id'])
                conference_edit_form = ConferenceEditForm(instance=event)
                event_option_set = event.eventoptions_set.all()
            except (Event.DoesNotExist, MultiValueDictKeyError):
                edit_action = 'blank'

    venue_list = Venue.objects.all().order_by('city', 'name')
    venue_form = VenueForm()
    conference_select_form = ConferenceSelectForm()
    context = {
        'edit_action': edit_action,
        'venue_form': venue_form,
        'venue_list': venue_list,
        'conference_select_form': conference_select_form,
        'conference_edit_form': conference_edit_form,
        'conference_option_form': conference_option_form,
        'event': event,
        'event_option_set': event_option_set,
    }
    return render(request, 'registration/conference.html', context)


@login_required
def index(request):
    return render(request, 'registration/index.html')


@login_required
def mass_mail(request):
    if request.method != 'POST':
        return redirect('/registration/')
    event = get_object_or_404(Event, pk=request.POST['event'])
    email_merge_fields = {
        'venue_details': request.POST['venue_details'].strip().replace('\r\n', '<br/>'),
        'event_registrar': request.POST['event_registrar'].strip().replace('\r\n', '<br/>'),
        'conference_name': request.POST['conference_name'].strip(),
        'conference_location': request.POST['conference_location'].strip(),
        'start_date': request.POST['start_date'].strip()
    }
    email_subject = 'Infonex - ' + request.POST['conference_name'].strip() + ' - '

    if request.POST['mass_mail_message'] == 'venue':
        email_body_path = os.path.join(
            BASE_DIR,
            'registration/static/registration/email_text/venue_details.txt'
        )
        if request.POST['room_rate'] not in ('', None):
            room_rate_text = '\nThe hotel is able to offer you a special ' + \
                             'guest room rate of '
            room_rate_text += request.POST['room_rate'].strip() + '. '
            if request.POST['room_rate_code'] not in ('', None):
                if request.POST['room_booking_phone'] not in ('', None):
                    room_rate_text += 'Please call ' + \
                        request.POST['room_booking_phone'].strip() + \
                        ' directly to book your room and quote ' + \
                        request.POST['room_rate_code'].strip() + '.<br/>'
                else:
                    room_rate_text += 'Please quote ' + \
                        request.POST['room_rate_code'].strip() + \
                        ' when booking your room.<br/>'
        else:
            room_rate_text = ''
        email_merge_fields['room_rate_text'] = room_rate_text
        email_subject += 'VENUE UPDATE'

    elif request.POST['mass_mail_message'] == 'docs':
        email_body_path = os.path.join(
            BASE_DIR,
            'registration/static/registration/email_text/doc_download.txt'
        )
        email_merge_fields['download_link'] = \
            request.POST['download_link'].strip()
        email_merge_fields['registration_start'] = \
            request.POST['registration_start'].strip()
        email_merge_fields['opening_remarks_time'] = \
            request.POST['opening_remarks_time'].strip()
        email_subject += 'DOCUMENT ACCESS'

    else:
        raise Http404('Invalid mail choice')
    with open(email_body_path) as f:
        email_body = f.read()
    email_body = email_body.format(**email_merge_fields)
    print(email_body)
    context = {
        'email_subject': email_subject,
        'email_body': email_body,
    }
    return render(request, 'registration/mass_mail.html', context)


@login_required
def new_delegate_search(request):
    """ Renders new_delegate_search page"""
    if request.method == 'POST':
        person_form = NewDelegateSearchForm(request.POST)
        if (request.POST['first_name'] == '' and
            request.POST['last_name'] == '' and
            request.POST['company'] == '' and
            request.POST['postal_code'] == '' and
            # TODO: This probably should be changed until event search
            # functionality is enabled
            request.POST['event'] == ''):

            past_customer_list = None
            crm_list = None
            search_entered = True
        else:
            filterargs = {
                'first_name__icontains': request.POST['first_name'],
                'last_name__icontains': request.POST['last_name'],
                'company__name__icontains': request.POST['company'],
                'company__postal_code__icontains': request.POST['postal_code'],
            }
            if request.POST['event'] != '':
                filterargs['regdetails__conference__pk'] = request.POST['event']
            past_customer_list = Registrants.objects.filter(**filterargs)

            crm_list = Person.objects.filter(
                Q(name__icontains=request.POST['first_name'].strip()) &
                Q(name__icontains=request.POST['last_name'].strip()),
                Q(company__icontains=request.POST['company'].strip())
            ).order_by('company', 'name')[:100]
            search_entered = True
    else:
        person_form = NewDelegateSearchForm()
        past_customer_list = None
        crm_list = None
        search_entered = None
    context = {
        'person_form': person_form,
        'past_customer_list': past_customer_list,
        'crm_list': crm_list,
        'search_entered': search_entered,
    }
    return render(request, 'registration/new_delegate_search.html', context)


#######################
# AJAX CALLS
#######################
@login_required
def add_event_option(request):
    """ ajax call to add options to a conference """
    conference_option_form = ConferenceOptionForm()
    event_option_set = None
    event = None
    if request.method == 'POST':
        try:
            event_id = request.POST['event_id']
        except MultiValueDictKeyError:
            event_id = None
        conference_option_form = ConferenceOptionForm(request.POST)
        if conference_option_form.is_valid():
            if not event_id:
                new_event_number = create_temp_conf_number()
                event = Event(
                    number=new_event_number,
                    title='Placeholder Event',
                    city='Placeholder City',
                    date_begins=timezone.now(),
                    registrar=request.user,
                    state_prov='ON',
                    created_by=request.user,
                    modified_by=request.user,
                )
                event.save()
            else:
                event = Event.objects.get(pk=event_id)
            option = EventOptions(
                event=event,
                name=request.POST['name'],
                startdate=request.POST['startdate'],
                enddate=request.POST['enddate'],
                primary=True if request.POST['primary'] == 'true' else False,
            )
            option.save()
            conference_option_form = ConferenceOptionForm()
            event_option_set = event.eventoptions_set.all()
        else:
            if event_id:
                event = Event.objects.get(pk=event_id)
                event_option_set = event.eventoptions_set.all()
    context = {
        'conference_option_form': conference_option_form,
        'event_option_set': event_option_set,
        'event': event,
    }
    return render(request, 'registration/addins/conference_options_panel.html',
                  context)


@login_required
def add_venue(request):
    """ AJAX Function to add new venue and refresh venue sidebar """
    errors = False
    venue_list = Venue.objects.all().order_by('city', 'name')
    if request.method == 'POST':
        venue_form = VenueForm(request.POST)
        errors = True
        if venue_form.is_valid():
            errors = False
            venue_form.save()
            venue_form = VenueForm()
    else:
        venue_form = VenueForm()
    context = {
        'venue_form': venue_form,
        'venue_list': venue_list,
        'errors': errors,
    }
    return render(request, 'registration/addins/venue_sidebar.html', context)


@login_required
def delete_event_option(request):
    """ ajax call to delete option from a conference """
    conference_option_form = ConferenceOptionForm()
    event_option_set = None
    event = None
    if request.method == 'POST':
        try:
            event_id = request.POST['event_id']
            event = Event.objects.get(pk=event_id)
        except (MultiValueDictKeyError, Event.DoesNotExist):
            event_id = None
        try:
            option_id = request.POST['option_id']
            option = EventOptions.objects.get(pk=option_id)
        except (EventOptions.DoesNotExist, MultiValueDictKeyError):
            option_id = None
        if event_id and option_id:
            option.delete()
            event_option_set = event.eventoptions_set.all()
    context = {
        'conference_option_form': conference_option_form,
        'event_option_set': event_option_set,
        'event': event,
    }
    return render(request, 'registration/addins/conference_options_panel.html',
                  context)


@login_required
def delete_temp_conf(request):
    """ ajax call to delete temporary conference and options """
    event = None
    if request.method == 'POST':
        try:
            event = Event.objects.get(pk=request.POST['event_id'])
        except (Event.DoesNotExist, MultiValueDictKeyError):
            event = None
        if event:
            event.delete()
    return HttpResponse('')


@login_required
def delete_venue(request):
    """ AJAX function to delete venue """
    venue_form = VenueForm()
    # TODO: Update this when filter function implemented
    venue_list = Venue.objects.all().order_by('city', 'name')
    if request.method == 'POST' and 'venue_id' in request.POST:
        try:
            venue = Venue.objects.get(id=request.POST['venue_id'])
        except (Venue.DoesNotExist, MultiValueDictKeyError):
            venue = None
        if venue:
            venue.delete()
    context = {
        'venue_form': venue_form,
        'venue_list': venue_list,
    }
    return render(request, 'registration/addins/venue_sidebar.html', context)


@login_required
def edit_venue(request):
    """ AJAX function to edit venue information in sidebar """
    venue = None
    venue_id = None
    template = 'registration/addins/venue_detail.html'
    if request.method == 'GET':
        venue_id = request.GET['id']
    elif request.method == 'POST':
        venue_id = request.POST['id']
    if venue_id:
        venue = Venue.objects.get(pk=venue_id)
        if request.method == 'GET':
            venue_form = VenueForm(instance=venue)
            template = 'registration/addins/venue_edit.html'
        elif request.method == 'POST':
            venue_form = VenueForm(request.POST, instance=venue)
            if venue_form.is_valid():
                venue_form.save()
            else:
                template = 'registration/addins/venue_edit.html'
        else:
            venue_form = VenueForm()

    context = {'venue': venue,
               'venue_form': venue_form,}
    return render(request, template, context)


@login_required
def filter_venue(request):
    """
    Dynamically filter list of venues based on city
    """
    venue_form = VenueForm()
    try:
        city_partial = request.GET['city_partial']
        venue_list = Venue.objects.filter(
            city__icontains=city_partial).order_by('city', 'name')
    except MultiValueDictKeyError:
        city_partial = None
        venue_list = Venue.objects.all().order_by('city', 'name')
    context = {
        'venue_form': venue_form,
        'venue_list': venue_list,
    }
    return render(request, 'registration/addins/venue_sidebar.html', context)


@login_required
def find_reg(request):
    """
    Ajax call to search for existing registrations;
    called from registration.js from registration_search.html
    """
    # variable instantiation
    failure_message = None
    matched_reg = None
    reg_pool = None
    num_events = 0
    if request.method != 'POST':
        return HttpResponse('')
    if request.POST['invoice_number'] != '':
        try:
            invoice = Invoice.objects.get(pk=request.POST['invoice_number'])
        except Invoice.DoesNotExist:
            invoice = None
            failure_message = 'Invoice Does Not Exist'
    else:
        invoice = None
    if request.POST['conf_id'] != '':
        try:
            conf = Event.objects.get(pk=request.POST['conf_id'])
            num_events = 1
        except Event.DoesNotExist:
            conf = None
            failure_message = 'Event Does Not Exist'
    else:
        conf = None
    first_name = request.POST['first_name']
    last_name = request.POST['last_name']
    company = request.POST['company']

    # conduct search
    if invoice:
        matched_reg = RegDetails.objects.get(invoice=invoice)

    elif conf:
        reg_pool = RegDetails.objects.filter(conference=conf)
        if reg_pool.count() == 0:
            failure_message == 'No registrations yet for that event'
        elif reg_pool.count() == 1 and not (first_name or last_name or company):
            matched_reg = reg_pool[0]

    if not matched_reg and not failure_message:
        reg_pool = RegDetails.objects.all() if reg_pool is None else reg_pool
        kwargs = {
            'registrant__first_name__icontains': first_name,
            'registrant__last_name__icontains': last_name,
            'registrant__company__name__icontains': company
        }
        reg_pool = reg_pool.filter(**kwargs)
        if reg_pool.count() == 0:
            failure_message = 'No registrations match your search criteria'
        elif reg_pool.count() == 1:
            matched_reg = reg_pool[0]

    # Case 1: Only one match
    if matched_reg:
        return JsonResponse({'reg_id': matched_reg.pk})
    # Case 2: No matches
    if failure_message:
        context = {'failure_message': failure_message}
    # Case 3: matches from only one or multiple conference
    else:
        reg_pool = reg_pool.order_by('-conference__date_begins',
                                     'registrant__last_name',
                                     'registrant__first_name')
        context = {'reg_match_list': reg_pool}
    return render(request,
                  'registration/index_panels/registration_search_results.html',
                  context)


@login_required
def get_registration_history(request):
    """AJAX call to get reg history for past delegate"""
    context = RequestContext(request)
    person_id = None
    person_regs = None
    if request.method == 'GET':
        person_id = request.GET['id']
    if person_id:
        person = Registrants.objects.get(pk=person_id)
        person_regs = person.regdetails_set.all()
    return render(request, 'registration/addins/reg_history.html',
        {'person_regs': person_regs})


@login_required
def index_panel(request):
    if 'panel' not in request.GET:
        return HttpResponse('')
    if request.GET['panel'] == 'admin-reports':
        response_url = 'registration/index_panels/admin_reports.html'
        conference_select_form = ConferenceSelectForm()
        conference_select_form.fields['event'].queryset = \
            Event.objects.all().order_by('-number')
        admin_report_options_form = AdminReportOptionsForm()
        context = {
            'conference_select_form': conference_select_form,
            'admin_report_options_form': admin_report_options_form,
        }
    elif request.GET['panel'] == 'reg-search':
        response_url = 'registration/index_panels/registration_search.html'
        delegate_search_form = NewDelegateSearchForm()
        # delegate_search_form.fields['event'].queryset = \
        #     Event.objects.all().order_by('-date_begins')
        context = {
            'delegate_search_form': delegate_search_form,
        }
    elif request.GET['panel'] == 'invoice-quick-search':
        invoice_number = request.GET.get('invoice_number', 0)
        response_url = 'registration/index_panels/invoice_details.html'
        try:
            invoice = Invoice.objects.get(id=invoice_number)
        except Invoice.DoesNotExist:
            invoice = None
        if invoice:
            taxes = invoice.pre_tax_price * invoice.gst_rate
            taxes += invoice.pre_tax_price * invoice.hst_rate
            taxes += (invoice.pre_tax_price * invoice.gst_rate +
                      invoice.pre_tax_price) * invoice.qst_rate
            total_invoice = invoice.pre_tax_price + taxes
        else:
            taxes = None
            total_invoice = None
        context = {
            'invoice': invoice,
            'taxes': taxes,
            'total_invoice': total_invoice,
        }
    elif request.GET['panel'] == 'sales-reports':
        response_url = 'registration/index_panels/sales_reports.html'
        select_form = SalesReportOptionsForm()
        context = {
            'select_form': select_form
        }
    elif request.GET['panel'] == 'event-revenue':
        response_url = 'registration/index_panels/event_revenue.html'
        conference_select_form = ConferenceSelectForm()
        conference_select_form.fields['event'].queryset = \
            Event.objects.all().order_by('-number')
        options_form = AdminReportOptionsForm()
        context = {
            'select_form': conference_select_form,
            'options_form': options_form
        }
    elif request.GET['panel'] == 'mass-mail':
        response_url = 'registration/index_panels/mass_mail.html'
        conference_select_form = ConferenceSelectForm()
        mass_mail_options_form = MassMailOptionsForm()
        context = {
            'conference_select_form': conference_select_form,
            'mass_mail_options_form': mass_mail_options_form,
        }

    else:
        raise Http404('Invalid panel name')

    return render(request, response_url, context)


@login_required
def mass_mail_details(request):
    message = request.GET.get('message', '')
    event_number = request.GET.get('event', None)
    event = get_object_or_404(Event, pk=event_number)

    registrar_string = 'Infonex Registration Department ' \
                       '416-971-4177' \
                       'register@infonex.ca'
    if event.registrar:
        registrar_string = ''
        registrar = event.registrar
        if registrar.first_name and registrar.last_name:
            registrar_string += registrar.first_name + ' ' + registrar.last_name
        else:
            registrar_string += 'Infonex Registration Department'
        if registrar.userprofile.phone:
            registrar_string += '\n' + registrar.userprofile.phone
        else:
            registrar_string += '\n416-971-4177'
        if registrar.email:
            registrar_string += '\n' + registrar.email
        else:
            registrar_string += '\nregister@infonex.ca'

    venue_details = ''
    booking_phone = ''
    if event.hotel:
        hotel = event.hotel
        if hotel.name:
            venue_details += hotel.name
        if hotel.address:
            venue_details += '\n' + hotel.address
        if hotel.city:
            venue_details += '\n' + hotel.city + ', ' + \
                hotel.state_prov + ' ' + hotel.postal_code
        if hotel.phone:
            venue_details += '\nHotel Phone: ' + hotel.phone
            booking_phone = hotel.phone
        if event.hotel.hotel_url:
            venue_details += '\nHotel Website: ' + hotel.hotel_url

    try:
        start_date = event.date_begins.strftime('%A, %-d %B, %Y')
    except AttributeError:
        start_date = None

    form_data = {
        'venue_details': venue_details,
        'event_registrar': registrar_string,
        'conference_name': event.title,
        'conference_location': event.city + ', ' + event.state_prov,
        'start_date': start_date,
        'room_booking_phone': booking_phone,
    }

    if message == 'venue':
        detail_panel = 'registration/index_panels/mass_mail_venue.html'
        merge_field_form = MailMergeDetailsForm(initial=form_data)

    elif message == 'docs':
        detail_panel = 'registration/index_panels/mass_mail_docs.html'
        download_link = 'http://www.infonex.ca/' + event.number + \
            '/download.shtml'
        form_data['download_link'] = download_link
        merge_field_form = MailMergeDetailsForm(initial=form_data)
    else:
        raise Http404('Invalid Message')

    context = {
        'merge_field_form': merge_field_form,
        'event': event,
    }
    return render(request, detail_panel, context)


@login_required
def save_conference_changes(request):
    """ AJAX call to save changes to conference being edited """
    event = None
    edit_action = None
    conference_option_form = ConferenceOptionForm()
    conference_edit_form = ConferenceEditForm()
    event_option_set = None
    if request.method == 'POST':
        try:
            event_id = request.POST['event_id']
            if event_id != 'new':
                event = Event.objects.get(pk=event_id)
            conference_edit_form = ConferenceEditForm(
                request.POST, instance=event
            )
            if conference_edit_form.is_valid():
                event = conference_edit_form.save()
                return HttpResponse(
                    '<div class="row text-center">' \
                    '<a href="/registration/new_delegate_search/" class="btn btn-primary">Create New Registration</a>' \
                    '</div>'
                )
        except (Event.DoesNotExist, MultiValueDictKeyError):
            conference_edit_form = ConferenceEditForm(request.POST)
    context = {
        'edit_action': edit_action,
        'conference_edit_form': conference_edit_form,
        'conference_option_form': conference_option_form,
        'event': event,
        'event_option_set': event_option_set,
    }
    return render(request, 'registration/addins/conference_edit_panel.html',
                  context)


@login_required
def search_dels(request):
    """
    AJAX call to get matching delegates and crm contacts for new_delegate_search
    """
    if request.method == 'POST':
        if (request.POST['first_name'] == '' and
            request.POST['last_name'] == '' and
            request.POST['company'] == '' and
            request.POST['postal_code'] == ''):

            past_customer_list = None
            crm_list = None
            search_entered = True
            conference = None
        else:
            filterargs = {
                'first_name__icontains': request.POST['first_name'],
                'last_name__icontains': request.POST['last_name'],
                'company__name__icontains': request.POST['company'],
                'company__postal_code__icontains': request.POST['postal_code'],
            }
            past_customer_list = Registrants.objects.filter(**filterargs)
            crm_list = Person.objects.filter(
                Q(name__icontains=request.POST['first_name']) &
                Q(name__icontains=request.POST['last_name']),
                Q(company__icontains=request.POST['company'])
            ).order_by('company', 'name')[:100]
            search_entered = True
            try:
                conference = Event.objects.get(pk=request.POST['event'])
            except (ValueError, Event.DoesNotExist):
                conference = None
    else:
        past_customer_list = None
        crm_list = None
        search_entered = None
        conference = None
    context = {
        'past_customer_list': past_customer_list,
        'crm_list': crm_list,
        'search_entered': search_entered,
        'conference': conference,
    }
    return render(request, 'registration/addins/match_del_list.html', context)


@login_required
def select_conference_to_edit(request):
    """ ajax call to select conference for editing """
    event = None
    edit_action = 'new'
    conference_edit_form = ConferenceEditForm()
    conference_option_form = ConferenceOptionForm()
    event_option_set = None
    if request.method == 'POST':
        if request.POST['edit_action'] != 'edit':
            edit_action = 'blank'
        else:
            try:
                event = Event.objects.get(pk=request.POST['conf_id'])
                edit_action = 'edit'
                conference_edit_form = ConferenceEditForm(instance=event)
                event_option_set = event.eventoptions_set.all()
            except (Event.DoesNotExist, MultiValueDictKeyError):
                edit_action = 'blank'
    context = {
        'edit_action': edit_action,
        'conference_edit_form': conference_edit_form,
        'conference_option_form': conference_option_form,
        'event': event,
        'event_option_set': event_option_set,
    }
    return render(request, 'registration/addins/conference_edit_panel.html',
                  context)


@login_required
def suggest_first_name(request):
    """
    Ajax call - returns json of top 25 first names (by number in db) that
    match entered string
    """
    query_term = request.GET.get('q', '')
    selects = Registrants.objects.filter(first_name__icontains=query_term) \
        .values('first_name').annotate(total=Count('first_name')) \
        .order_by('-total')[:25]
    results = []
    for select in selects:
        select_json = {}
        select_json['identifier'] = select['first_name']
        results.append(select_json)
    data = json.dumps(results)
    mimetype = 'applications/json'
    return HttpResponse(data, mimetype)


@login_required
def suggest_last_name(request):
    """
    Ajax call - returns json of top 25 first names (by number in db) that
    match entered string
    """
    query_term = request.GET.get('q', '')
    selects = Registrants.objects.filter(last_name__icontains=query_term) \
        .values('last_name').annotate(total=Count('last_name')) \
        .order_by('-total')[:25]
    results = []
    for select in selects:
        select_json = {}
        select_json['identifier'] = select['last_name']
        results.append(select_json)
    data = json.dumps(results)
    mimetype = 'applications/json'
    return HttpResponse(data, mimetype)


@login_required
def unfilter_venue(request):
    """
    AJAX Call
    """
    venue_form = VenueForm()
    venue_list = Venue.objects.all().order_by('city', 'name')
    context = {
        'venue_form': venue_form,
        'venue_list': venue_list,
    }
    return render(request, 'registration/addins/venue_sidebar.html', context)


@login_required
def update_conference_choices(request):
    """ ajax call to update conference select dropdown on conference.html """
    if 'qs' in request.GET:
        conference_select_form = ConferenceSelectForm()
        if request.GET['qs'] == 'all':
            conference_select_form.fields['event'].queryset = \
                Event.objects.all().order_by('-number')
    else:
        conference_select_form = ConferenceSelectForm()
    return render(request,
                  'registration/addins/conference_select_dropdown.html',
                  {'conference_select_form': conference_select_form})


@login_required
def update_venue_choices(request):
    """ ajax call to update venue selection choices dropdown when new venue added """
    conference_edit_form = ConferenceEditForm()
    return render(request, 'registration/addins/conference_venue_choices.html',
                  {'conference_edit_form': conference_edit_form})


############################
# GRAPHICS/DOCUMENTS
############################
@login_required
def get_admin_reports(request):
    # Pull relevant info out of GET request
    if 'event' not in request.GET:
        raise Http404('Event not specified')
    event = get_object_or_404(Event, pk=request.GET['event'])
    sort = request.GET.get('sort', 'company')
    if sort not in ('company', 'name', 'title'):
        sort = 'company'
    destination = request.GET.get('destination', 'inline')
    if destination not in ('attachment', 'inline'):
        destination = 'attachment'
    doc_format = request.GET.get('doc_format', 'pdf')
    if doc_format not in ('pdf', 'csv', 'xlsx'):
        doc_format = 'pdf'
    report_type = request.GET.get('report', '')
    sort_orders = {
        'name': ['registrant__last_name',
                 'registrant__first_name',
                 'registrant__company__name'],
        'title': ['registrant__title',
                  'registrant__company__name',
                  'registrant__last_name',
                  'registrant__first_name'],
        'company': ['registrant__company__name',
                    'registrant__title',
                    'registrant__last_name',
                    'registrant__first_name']
    }

    if doc_format == 'pdf':
        buffr = BytesIO()
        report = ConferenceReportPdf(buffr, event, sort)
        response = HttpResponse(content_type='application/pdf')
        if report_type == 'Delegate':
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_delegate_list()
        elif report_type == 'NoName':
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_no_name_list()
        elif report_type == 'Registration':
            file_details = destination + '; filename="registration_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_registration_list()
        elif report_type == 'Phone':
            file_details = destination + '; filename="phone_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_phone_list()
        elif report_type == 'Onsite':
            file_details = destination + '; filename="onsite_delegate_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_onsite_list()
        elif report_type == 'Unpaid':
            file_details = destination + '; filename="unpaid_delegate_list_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_unpaid_list()
        elif report_type == 'CE':
            file_details = destination + '; filename="CE_signin_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.generate_ce_sign_in_sheet()
        elif report_type == 'Badges':
            file_details = destination + '; filename="badges_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.badges()
        elif report_type == 'Counts':
            file_details = destination + '; filename="count_report_' + \
                str(event.number) + '.pdf"'
            response['Content-Disposition'] = file_details
            pdf = report.delegate_count()
        else:  # Invalid report type
            buffr.close()
            raise Http404('Invalid report type')
        buffr.close()
        response.write(pdf)
        return response
    elif doc_format == 'csv':
        response = HttpResponse(content_type='text/csv')

        if report_type == 'Delegate':
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.csv"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(
                *sort_orders[sort]
            )
            writer = csv.writer(response)
            writer.writerow(['Name', 'Title', 'Company'])
            for record in registration_qs:
                registrant = record.registrant
                writer.writerow([
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name
                ])
            return response

        elif report_type == 'NoName':
            sort = 'company' if sort == 'name' else sort
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.csv"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(
                *sort_orders[sort]
            )
            writer = csv.writer(response)
            writer.writerow(['Title', 'Company'])
            for record in registration_qs:
                registrant = record.registrant
                writer.writerow([
                    registrant.title,
                    registrant.company.name
                ])
            return response

        elif report_type == 'Registration':
            file_details = destination + '; filename="registration_list_' + \
                str(event.number) + '.csv"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=NO_CONFIRMATION_VALUES
            ).order_by(*sort_orders[sort])
            writer = csv.writer(response)
            writer.writerow(['Status', 'RegDate', 'InvoiceNumber',
                             'PreTaxPrice', 'Name', 'Title', 'Company',
                             'City', 'StateProv', 'SalesCredit'])
            for record in registration_qs:
                registrant = record.registrant
                try:
                    invoice_num = record.invoice.pk
                    pre_tax_price = record.invoice.pre_tax_price
                    if record.invoice.sales_credit.first_name or \
                        record.invoice.sales_credit.last_name:
                        sales_credit = record.invoice.sales_credit.first_name \
                            + ' ' + record.invoice.sales_credit.last_name
                    else:
                        sales_credit = record.invoice.sales_credit.username
                except Invoice.DoesNotExist:
                    invoice_num = ''
                    pre_tax_price = 0
                    sales_credit = ''
                writer.writerow([
                    record.registration_status,
                    record.register_date,
                    invoice_num,
                    pre_tax_price,
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.company.city,
                    registrant.company.state_prov,
                    sales_credit,
                ])
            return response

        elif report_type == 'Phone':
            file_details = destination + '; filename="phone_list_' + \
                str(event.number) + '.csv"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(*sort_orders[sort])
            writer = csv.writer(response)
            writer.writerow(['name', 'title', 'company', 'email', 'phone',
                             'city', 'stateProv', 'nameForLetters'])
            for record in registration_qs:
                registrant = record.registrant
                letter_name = ''
                if registrant.salutation:
                    letter_name += registrant.salutation + ' '
                else:
                    letter_name += registrant.first_name + ' '
                letter_name += registrant.last_name
                writer.writerow([
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.email1,
                    registrant.phone1,
                    registrant.company.city,
                    registrant.company.state_prov,
                    letter_name,
                ])
            return response

        elif report_type == 'Speaker':
            file_details = destination + '; filename="speaker_list_' + \
                str(event.number) + '.csv"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event, registration_status='K'
            ).order_by(*sort_orders[sort])
            writer = csv.writer(response)
            writer.writerow(['salutation', 'firstName', 'lastName', 'name',
                             'nameForLetters', 'title', 'company', 'address1',
                             'address2', 'city', 'stateProv', 'postalCode',
                             'country', 'phone', 'email'])
            for record in registration_qs:
                registrant = record.registrant
                letter_name = ''
                if registrant.salutation:
                    letter_name += registrant.salutation + ' '
                else:
                    letter_name += registrant.first_name + ' '
                letter_name += registrant.last_name
                writer.writerow([
                    registrant.salutation,
                    registrant.first_name,
                    registrant.last_name,
                    registrant.first_name + ' ' + registrant.last_name,
                    letter_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.company.address1,
                    registrant.company.address2,
                    registrant.company.city,
                    registrant.company.state_prov,
                    registrant.company.postal_code,
                    registrant.company.country,
                    registrant.phone1,
                    registrant.email1,
                ])
            return response

        raise Http404('Invalid Report Choice for CSV Export')

    elif doc_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb = Workbook()
        ws = wb.active

        if report_type == 'Delegate':
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.xlsx"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(
                *sort_orders[sort]
            )
            ws.append(['Name', 'Title', 'Company'])
            for record in registration_qs:
                registrant = record.registrant
                ws.append([
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name
                ])
            wb.save(response)
            return response

        elif report_type == 'NoName':
            sort = 'company' if sort == 'name' else sort
            file_details = destination + '; filename="delegate_list_' + \
                str(event.number) + '.xlsx"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(
                *sort_orders[sort]
            )
            ws.append(['Title', 'Company'])
            for record in registration_qs:
                registrant = record.registrant
                ws.append([
                    registrant.title,
                    registrant.company.name
                ])
            wb.save(response)
            return response

        elif report_type == 'Registration':
            file_details = destination + '; filename="registration_list_' + \
                str(event.number) + '.xlsx"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=NO_CONFIRMATION_VALUES
            ).order_by(*sort_orders[sort])
            ws.append(['Status', 'RegDate', 'InvoiceNumber', 'PreTaxPrice',
                       'Name', 'Title', 'Company', 'City', 'StateProv',
                       'SalesCredit'])
            for record in registration_qs:
                registrant = record.registrant
                try:
                    invoice_num = record.invoice.pk
                    pre_tax_price = record.invoice.pre_tax_price
                    if record.invoice.sales_credit.first_name or \
                        record.invoice.sales_credit.last_name:
                        sales_credit = record.invoice.sales_credit.first_name \
                            + ' ' + record.invoice.sales_credit.last_name
                    else:
                        sales_credit = record.invoice.sales_credit.username

                except Invoice.DoesNotExist:
                    invoice_num = ''
                    pre_tax_price = 0
                    sales_credit = ''
                ws.append([
                    record.get_registration_status_display(),
                    record.register_date,
                    invoice_num,
                    pre_tax_price,
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.company.city,
                    registrant.company.state_prov,
                    sales_credit,
                ])
            wb.save(response)
            return response

        elif report_type == 'Phone':
            file_details = destination + '; filename="phone_list_' + \
                str(event.number) + '.xlsx"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event
            ).exclude(
                registration_status__in=CXL_VALUES
            ).order_by(*sort_orders[sort])
            ws.append(['name', 'title', 'company', 'email', 'phone', 'city',
                       'stateProv', 'nameForLetters'])
            for record in registration_qs:
                registrant = record.registrant
                letter_name = ''
                if registrant.salutation:
                    letter_name += registrant.salutation + ' '
                else:
                    letter_name += registrant.first_name + ' '
                letter_name += registrant.last_name
                ws.append([
                    registrant.first_name + ' ' + registrant.last_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.email1,
                    registrant.phone1,
                    registrant.company.city,
                    registrant.company.state_prov,
                    letter_name,
                ])
            wb.save(response)
            return response

        elif report_type == 'Speaker':
            file_details = destination + '; filename="speaker_list_' + \
                str(event.number) + '.xlsx"'
            response['Content-Disposition'] = file_details
            registration_qs = RegDetails.objects.filter(
                conference=event, registration_status='K'
            ).order_by(*sort_orders[sort])
            ws.append(['salutation', 'firstName', 'lastName', 'name',
                       'nameForLetters', 'title', 'company', 'address1',
                       'address2', 'city', 'stateProv', 'postalCode',
                       'country', 'phone', 'email'])
            for record in registration_qs:
                registrant = record.registrant
                letter_name = ''
                if registrant.salutation:
                    letter_name += registrant.salutation + ' '
                else:
                    letter_name += registrant.first_name + ' '
                letter_name += registrant.last_name
                ws.append([
                    registrant.salutation,
                    registrant.first_name,
                    registrant.last_name,
                    registrant.first_name + ' ' + registrant.last_name,
                    letter_name,
                    registrant.title,
                    registrant.company.name,
                    registrant.company.address1,
                    registrant.company.address2,
                    registrant.company.city,
                    registrant.company.state_prov,
                    registrant.company.postal_code,
                    registrant.company.country,
                    registrant.phone1,
                    registrant.email1,
                ])
            wb.save(response)
            return response

        raise Http404('Invalid Report Choice for Excel Export')

    else:
        raise Http404('Invalid document format')


@login_required
def event_revenue(request):
    if 'event' not in request.GET:
        raise Http404('Event not specified')
    event = get_object_or_404(Event, pk=request.GET['event'])
    destination = request.GET.get('destination', 'attachment')
    if destination not in ('attachment', 'inline'):
        destination = 'attachment'
    doc_format = request.GET.get('doc_format', 'pdf')
    if doc_format not in ('pdf', 'csv', 'xlsx'):
        doc_format = 'pdf'
    revenue_qs = Invoice.objects.filter(
        reg_details__conference=event
    ).exclude(
        reg_details__registration_status__in=NON_INVOICE_VALUES
    ).order_by('pk')

    if doc_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb = Workbook()
        ws = wb.active
        ws.name = 'Event Revenue for ' + event.number
        file_details = destination + '; filename="revenue_report_' + \
            str(event.number) + '.xlsx"'
        response['Content-Disposition'] = file_details
        ws.append(['Conference Revenue Report - ' + event.number + ' - ' +
                   event.title])
        format_sales_report_header_row(ws, 1)
        ws.append(['InvoiceNumber',
                   'Company',
                   'Sales Credit',
                   'Pre-Tax Price',
                   'CDN Equivalent'])
        format_sales_report_header_row(ws, 2)
        num_rows = 2
        for record in revenue_qs:
            if record.sales_credit.first_name or record.sales_credit.last_name:
                sales_credit = record.sales_credit.first_name + ' ' + \
                    record.sales_credit.last_name
            else:
                sales_credit = record.sales_credit.username
            ws.append([record.pk,
                       record.reg_details.registrant.company.name,
                       sales_credit,
                       record.pre_tax_price,
                       record.pre_tax_price * record.fx_conversion_rate])
            num_rows += 1
        ws.append(['Totals',
                   '',
                   '',
                   '=SUM(D3:D' + str(num_rows) + ')'
                   '=SUM(E3:E' + str(num_rows) + ')'])
        format_sales_report_header_row(ws, num_rows + 1)
        wb.save(response)
        return response

    elif doc_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        file_details = destination + '; filename="revenue_report_' + \
            str(event.number) + '.csv"'
        response['Content-Disposition'] = file_details
        writer = csv.writer(response)
        writer.writerow(['InvoiceNumber',
                         'Company',
                         'Sales Credit',
                         'Pre-Tax Price',
                         'CDN Equivalent'])
        for record in revenue_qs:
            if record.sales_credit.first_name or record.sales_credit.last_name:
                sales_credit = record.sales_credit.first_name + ' ' + \
                    record.sales_credit.last_name
            else:
                sales_credit = record.sales_credit.username
            writer.writerow([record.pk,
                             record.reg_details.registrant.company.name,
                             sales_credit,
                             record.pre_tax_price,
                             record.pre_tax_price * record.fx_conversion_rate])
        writer.writerow(['',
                         '',
                         '',
                         revenue_qs.aggregate(Sum(
                             'pre_tax_price'
                         ))['pre_tax_price__sum']])
        return response

    else:  # doc_format = 'pdf'
        buffr = BytesIO()
        report = ConferenceReportPdf(buffr, event)
        response = HttpResponse(content_type='application/pdf')
        file_details = destination + '; filename="delegate_list_' + \
            str(event.number) + '.pdf"'
        response['Content-Disposition'] = file_details
        pdf = report.event_revenue(revenue_qs)
        buffr.close()
        response.write(pdf)
        return response


@login_required
def get_sales_reports(request):
    if 'report_date_month' not in request.GET or \
        'report_date_year' not in request.GET:
        raise Http404('Date not specified')
    sales_form = SalesReportOptionsForm(request.GET)
    if not sales_form.is_valid():
        raise Http404('Invalid Date Information')
    report_date = sales_form.cleaned_data.get('report_date')
    report_year = report_date.year
    report_month = report_date.month
    month_text = report_date.strftime('%B')

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="sales_data_' + \
        month_text + '_' + str(report_year) + '.xlsx"'
    wb = Workbook()
    column_names = ['Sales Credit',
                    'Invoice Number',
                    'Event Number',
                    'Event Name',
                    'Registration Status',
                    'Delegate Name',
                    'Company',
                    'Register Date',
                    'Payment Date',
                    'Payment Method',
                    'Cancellation Date',
                    'Pre-Tax Price',
                    'Billing Currency',
                    'FX Conversion Rate',
                    'CDN Equivalent']

    # Save monthly booking info
    ws1 = wb.active
    ws1.title = 'Monthly Sales for ' + month_text
    ws1.append(column_names)
    format_sales_report_header_row(ws1)
    monthly_bookings = RegDetails.objects.filter(
        register_date__year=report_year, register_date__month=report_month
    ).exclude(registration_status__in=NON_INVOICE_VALUES).order_by(
        'invoice__sales_credit', 'register_date', 'date_created'
    )
    for record in monthly_bookings:
        ws1.append(sales_report_row(record))

    # Save paid information
    ws2 = wb.create_sheet('Payments Received in ' + month_text)
    ws2.append(column_names)
    format_sales_report_header_row(ws2)
    monthly_payments = RegDetails.objects.filter(
        invoice__payment_date__year=report_year,
        invoice__payment_date__month=report_month
    ).order_by('invoice__sales_credit',
               'invoice__payment_date',
               'register_date',
               'date_created'
    )
    for record in monthly_payments:
        ws2.append(sales_report_row(record))

    # Save Unpaid List
    ws3 = wb.create_sheet('Unpaid (as of %s)' % str(datetime.date.today()))
    ws3.append(column_names)
    format_sales_report_header_row(ws3)
    unpaid_regs = RegDetails.objects.filter(
        registration_status__in=UNPAID_STATUS_VALUES
    ).order_by(
        'invoice__sales_credit', 'register_date', 'date_created'
    )
    for record in unpaid_regs:
        ws3.append(sales_report_row(record))

    # Save cancellations
    ws4 = wb.create_sheet('Cancellations in ' + month_text)
    ws4.append(column_names)
    format_sales_report_header_row(ws4)
    monthly_cancellations = RegDetails.objects.filter(
        cancellation_date__year=report_year,
        cancellation_date__month=report_month
    ).exclude(registration_status__in=NON_INVOICE_VALUES).order_by(
        'invoice__sales_credit', 'cancellation_date', 'register_date'
    )
    for record in monthly_cancellations:
        ws4.append(sales_report_row(record))

    wb.save(response)
    return response
